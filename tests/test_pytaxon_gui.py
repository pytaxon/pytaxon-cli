import pytest
from unittest.mock import MagicMock, patch
import tkinter as tk
from tkinter import filedialog
from openpyxl import Workbook
from pytaxon.pytaxon_gui import (
    open_file,
    clear_frame,
    run_pytaxon_correct,
    clear_treeviews,
    run_pytaxon,
    load_spreadsheet,
)

@pytest.fixture
def app_with_widgets():
    """Cria uma aplicação com widgets básicos para teste."""
    root = tk.Tk()
    frame = tk.Frame(root)
    frame.pack()

    entry = tk.Entry(root)
    entry.pack()

    textbox = tk.Text(root)
    textbox.pack()

    tree = tk.ttk.Treeview(root)
    tree.pack()

    yield root, frame, entry, textbox, tree
    root.destroy()

def test_open_file(app_with_widgets):
    """Teste para a função open_file."""
    root, _, entry, _, _ = app_with_widgets

    with patch("tkinter.filedialog.askopenfilename") as mock_file_dialog:
        mock_file_dialog.return_value = "Modified Uropygi Collection.xlsx"
        open_file(entry)
        assert entry.get() == "Modified Uropygi Collection.xlsx"

def test_clear_frame(app_with_widgets):
    """Teste para a função clear_frame."""
    root, frame, _, _, _ = app_with_widgets

    tk.Label(frame, text="Test Label").pack()
    assert len(frame.winfo_children()) > 0

    clear_frame(frame)
    assert len(frame.winfo_children()) == 0

def test_load_spreadsheet(app_with_widgets):
    """Teste para a função load_spreadsheet."""
    _, _, entry, _, tree = app_with_widgets

    wb = Workbook()
    sheet = wb.active
    sheet.append(["Line", "Column1", "Column2"])
    sheet.append([1, "Test1", "Test2"])
    test_file_path = "Modified Uropygi Collection.xlsx"
    wb.save(test_file_path)

    with patch("openpyxl.load_workbook") as mock_load_workbook:
        mock_load_workbook.return_value = wb
        loaded_sheet, loaded_path = load_spreadsheet(test_file_path)
        assert loaded_path == test_file_path
        assert loaded_sheet.title == sheet.title
