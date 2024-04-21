import os
from collections import Counter

import customtkinter
import customtkinter as ctk
from CTkMessagebox import CTkMessagebox
from tkinter import filedialog, ttk, Toplevel, Entry, Button, Label
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import matplotlib.pyplot as plt
from PIL import Image, ImageTk
import openpyxl
import subprocess
from openpyxl import load_workbook
from ttkthemes import ThemedTk

from pytaxon import Pytaxon


def open_file(entry_widget):
    filetypes = [("Excel files", "*.xlsx"), ("CSV files", "*.csv"), ("All files", "*.*")]
    filepath = filedialog.askopenfilename(filetypes=filetypes)
    if filepath:
        entry_widget.delete(0, ctk.END)
        entry_widget.insert(0, filepath)

def clear_frame(frame):
    for widget in frame.winfo_children():
        widget.destroy()

def run_pytaxon_correct(input_entry, corrected_spreadsheet_entry):
    textbox.delete("1.0", "end")

    try:
        pt.update_original_spreadsheet(input_entry, str(path_to_correct_spreadsheet), corrected_spreadsheet_entry)
        textbox.insert('end', "Pytaxon correction has been run successfully.\n")

        clear_treeviews()  # Limpa as visualizações após a execução bem-sucedida

        # Limpa o campo corrected_spreadsheet_entry após a execução bem-sucedida
        # corrected_spreadsheet_entry.delete(0, ctk.END)
        clear_frame(frame_a)
        clear_frame(frame_b)
        clear_frame(frame_c)


    except subprocess.CalledProcessError as e:
        #CTkMessagebox(title="Error", message=f"An error occurred while running Pytaxon correction: {e}", icon="cancel")
        textbox.insert('end', f"An error occurred while running Pytaxon correction: {e}\n")
    except Exception as e:
        textbox.insert('end', f"An unexpected error occurred: {e}\n")


# Adicione esta nova função para limpar os Treeviews
def clear_treeviews():
    global tree
    for item in tree.get_children():
        tree.delete(item)
    tree['columns'] = []

def run_pytaxon(input_path, source_id, check_spreadsheet_name):
    conteudo_textbox = textbox.get("1.0", "end-1c")
    if conteudo_textbox.strip():
        clear_treeviews()
        textbox.delete("1.0", "end")
        clear_frame(frame_a)
        clear_frame(frame_b)
        clear_frame(frame_c)

    columns = entry_columns.get()

    try:
        global pt 
        pt = Pytaxon(source_id)
        pt.read_spreadshet(input_path)
        pt.read_columns(columns)
        pt.check_species_and_lineage()
        global path_to_correct_spreadsheet
        path_to_correct_spreadsheet = pt.create_to_correct_spreadsheet(check_spreadsheet_name)

        log_file_path = "spreadsheet_log.txt"

        # Verificar se o arquivo de log existe e contém a mensagem específica
        if os.path.exists(log_file_path):
            with open(log_file_path, 'r') as file:
                content = file.read().strip()
                if content == 'No errors in spreadsheet':
                    textbox.insert('end', "No errors in spreadsheet.\n")
                    os.remove(log_file_path)  # Apagar o arquivo de log
                    clear_treeviews()
                    return

        textbox.insert('end', "Pytaxon has been run successfully.\n")
        load_spreadsheet(input_path, source_id)
        calculate_statistics(input_path)
    except subprocess.CalledProcessError as e:
        #CTkMessagebox(title="Error", message=f"An error occurred while running Pytaxon: {e}", icon="cancel")
        textbox.insert('end', f"An error occurred while running Pytaxon: {e}\n")
    except Exception as e:
        textbox.insert('end', f"An unexpected error occurred: {e}\n")


def show_id_info():
    CTkMessagebox(message="IDs Data Sources:\n"
                          "1 - Catalogue of Life Checklist\n"
                          "4 - NBCI\n"
                          "11 - GBIF\n"
                          "180 - iNaturalist Taxonomy", option_1="Ok")

def load_spreadsheet(file_path, source_id=None):
    try:
        workbook = load_workbook(filename=file_path, data_only=True)
        sheet = workbook.active
        user_columns = entry_columns.get().split(',')

        headers = ["Line"] + [cell.value for cell in sheet[1] if cell.value in user_columns]
        data = [[idx + 2] + [row[idx].value for idx, cell in enumerate(sheet[1]) if cell.value in headers[1:]]
                for idx, row in enumerate(sheet.iter_rows(min_row=2))]


        if path_to_correct_spreadsheet and source_id:
            load_spreadsheet_additional(path_to_correct_spreadsheet, tree, source_id)
    except Exception as e:
        CTkMessagebox(title="Error", message=f"Erro ao carregar a planilha: {e}", icon="cancel")

    return sheet, file_path


def load_spreadsheet_additional(file_path, treeview, source_id):
    workbook_spreadsheet = load_workbook(filename=file_path, data_only=False)
    sheet_spreadsheet = workbook_spreadsheet.active

    # Mapeia os cabeçalhos para seus índices
    headers = [cell.value for cell in sheet_spreadsheet[1]]

    # Define as colunas que desejamos exibir no Treeview baseadas na fonte de dados
    if source_id == "1":  # Catalogue of Life Checklist
        id_column = 'COL ID Source'
    elif source_id == "4":  # NBCI
        id_column = 'NCBI ID Source'
    elif source_id == "11":  # GBIF
        id_column = 'GBIF ID Source'
    elif source_id == "180":  # iNaturalist Taxonomy
        id_column = 'INAT ID Source'
    else:
        id_column = None

    desired_columns = ['Error Line', 'Rank', 'Wrong Name', 'Suggested Name', id_column, 'Change'] if id_column else [
        'Error Line', 'Rank', 'Wrong Name', 'Suggested Name', 'Change']

    treeview['columns'] = desired_columns
    treeview['show'] = 'headings'

    # Configura o cabeçalho e a largura das colunas no treeview
    for col in desired_columns:
        treeview.heading(col, text=col)
        treeview.column(col, width=100, anchor='center')

    # Insere os dados nas colunas do treeview
    for row in sheet_spreadsheet.iter_rows(min_row=2):
        row_data = [(cell.value if cell.value is not None else "") for cell in row]
        if id_column:
            id_column_index = headers.index(id_column) if id_column in headers else None

            # Se o índice for válido e a célula contiver uma fórmula de hiperlink
            if id_column_index is not None and isinstance(row_data[id_column_index], str):
                row_data[id_column_index] = extract_url(row_data[id_column_index])

        # Reduzir row_data para apenas as colunas desejadas
        row_data = [row_data[headers.index(col)] for col in desired_columns]
        treeview.insert('', 'end', values=row_data)


def extract_url(hyperlink_formula):
    # Extrai a URL da fórmula do hiperlink Excel
    if hyperlink_formula and '=HYPERLINK("' in hyperlink_formula:
        # Extrai a URL entre as aspas duplas
        url_start = hyperlink_formula.find('"') + 1
        url_end = hyperlink_formula.find('";', url_start)
        return hyperlink_formula[url_start:url_end]
    return hyperlink_formula  # Retorna o valor original se não for um hiperlink


def on_double_click(event, treeview, filepath):
    col_id = treeview.identify_column(event.x)
    col_name = treeview.heading(col_id, 'text')

    item = treeview.selection()[0]
    row_index = treeview.index(item) + 2
    current_value = treeview.item(item, 'values')[treeview['columns'].index(col_name)]

    popup = Toplevel()
    popup.geometry("200x100+690+400")
    entry = Entry(popup)
    entry.insert(0, current_value)
    entry.pack()
    entry.focus_set()

    def save_new_value(entry, row_index, col_name, item, treeview, filepath):
        new_value = entry.get()

        workbook = load_workbook(filename=filepath)
        worksheet = workbook.active  # Assume que a planilha ativa é a correta

        col_letter = None
        for cell in worksheet[1]:  # Procura pelo cabeçalho correto para encontrar a letra da coluna
            if cell.value == col_name:
                col_letter = cell.column_letter
                cell_ref = f"{col_letter}{row_index}"
                worksheet[cell_ref].value = new_value
                print(f"Valor '{new_value}' salvo na célula '{cell_ref}'")
                break
        else:
            print(f"Não foi possível encontrar a coluna '{col_name}' na planilha.")
            return

        workbook.save(filepath)
        workbook.close()
        treeview.set(item, column=col_name, value=new_value)
        print("Treeview atualizado.")
        popup.destroy()

    button = Button(popup, text="Save", command=lambda: save_new_value(
        entry, row_index, col_name, item, treeview, filepath))
    button.pack()


def calculate_statistics(entry_input):
    # Caminhos dos arquivos
    user_spreadsheet_path = entry_input  # Caminho para a planilha do usuário

    # Abrindo as planilhas
    user_wb = openpyxl.load_workbook(user_spreadsheet_path)
    check_wb = openpyxl.load_workbook(path_to_correct_spreadsheet)

    # Selecionando as primeiras abas das planilhas
    user_sheet = user_wb.active
    check_sheet = check_wb.active

    # Lendo dados da planilha do usuário
    user_data = [row[:8] for row in user_sheet.iter_rows(min_row=2, values_only=True)]

    # Calculando a quantidade total de ocorrências
    total_occurrences = len(user_data)
    taxon_count = sum(1 for occurrence in user_data for cell in occurrence if cell is not None)

    # Lendo dados da planilha Checagem
    check_data = [row[:6] for row in check_sheet.iter_rows(min_row=2, values_only=True)]

    # Calculando a quantidade de ocorrências com erro
    unique_error_lines = set(row[0] for row in check_data)
    occurrences_with_error = len(unique_error_lines)

    # Verificação de divisão por zero
    if total_occurrences > 0:
        percentage_occurrences_with_error = (occurrences_with_error / total_occurrences) * 100
    else:
        percentage_occurrences_with_error = 0

    if taxon_count > 0:
        percentage_taxons_with_error = (occurrences_with_error / taxon_count) * 100
    else:
        percentage_taxons_with_error = 0

    # Contando erros por táxon
    taxon_errors = Counter(row[2] for row in check_data)

    # Encontrando os táxons com mais erros
    top_taxon_errors = taxon_errors.most_common(3)

    # Atualizar o dashboard
    create_dashboard(frame_a, frame_b, frame_c, total_occurrences, percentage_occurrences_with_error, taxon_count, percentage_taxons_with_error, top_taxon_errors)


def add_bar_graph(frame, top_taxon_errors):
    # Atualiza o frame para obter as dimensões corretas
    frame.update_idletasks()
    frame_width = frame.winfo_width()
    frame_height = frame.winfo_height()

    # Limitando os dados aos três principais táxons
    top_taxon_errors = top_taxon_errors[:3]

    # Dados para o gráfico de barras atualizados para os três principais táxons
    taxa = [taxon[0] for taxon in top_taxon_errors]  # Nomes dos táxons
    errors = [taxon[1] for taxon in top_taxon_errors]  # Contagem de erros

    # Cria a figura e o gráfico de barras
    fig, ax = plt.subplots(figsize=(frame_width / 100, frame_height / 100))
    fig.patch.set_facecolor('none')  # Torna o fundo da figura transparente
    ax.set_facecolor('none')  # Torna o fundo do eixo transparente

    # Define as posições das barras para que fiquem mais próximas
    y_pos_offset = 0.3
    y_pos = [i * y_pos_offset for i in range(len(taxa))]
    bars = ax.barh(y_pos, errors, height=0.1, color='#FF8C00')
    ax.set_yticks(y_pos)
    ax.set_yticklabels(['' for _ in taxa])  # Limpa os rótulos dos eixos y temporariamente

    # Ajusta os limites do eixo y para corresponder aos novos valores de y_pos
    ax.set_ylim(min(y_pos) - y_pos_offset, max(y_pos) + y_pos_offset)
    ax.invert_yaxis()  # Inverte o eixo y para que a contagem comece de cima para baixo

    # Personalização dos rótulos dos táxons
    for bar, taxon in zip(bars, taxa):
        ax.text(bar.get_x() + bar.get_width() / 2, bar.get_y() - y_pos_offset / 2,
                taxon,
                ha='left', va='center', color='white', weight='bold')

    # Personalização dos textos das barras
    for bar in bars:
        ax.text(bar.get_width(), bar.get_y() + bar.get_height() / 2,
                f' {bar.get_width()}',
                va='center', color='white', weight='bold')

    # Remove os eixos para um visual mais limpo
    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)
    ax.spines['bottom'].set_visible(False)
    ax.spines['left'].set_visible(False)
    ax.xaxis.set_visible(False)  # Esconde os valores no eixo x
    ax.yaxis.set_ticks_position('none')  # Esconde os ticks do eixo y

    # Configura o fundo do widget canvas como transparente e sem borda
    canvas = FigureCanvasTkAgg(fig, master=frame)
    canvas.draw()
    canvas_widget = canvas.get_tk_widget()
    canvas_widget.config(bg='#004C70', highlightthickness=0)  # Configura o widget para ser transparente
    canvas_widget.pack(fill='both', expand=True)


def add_pie_chart(frame, errors_percentage):
    # Certifique-se de que a porcentagem de erros está entre 0 e 100
    errors_percentage = max(0, min(errors_percentage, 100))
    no_errors_percentage = 100 - errors_percentage

    # Atualiza o frame para obter as dimensões corretas
    frame.update_idletasks()
    frame_width = frame.winfo_width()
    frame_height = frame.winfo_height()
    figsize_width = frame_width / 100
    figsize_height = frame_height / 100

    # Cria a figura e o gráfico de pizza com o tamanho padrão
    fig, ax = plt.subplots(figsize=(figsize_width, figsize_height))
    ax.pie([errors_percentage, no_errors_percentage], labels=['Errors', 'No Errors'],
           startangle=140, textprops={'color': 'white'},
           autopct='%1.1f%%')

    # Define a cor de fundo do gráfico
    fig.set_facecolor('#004C70')
    ax.set_facecolor('#004C70')

    # Adiciona o gráfico ao frame usando FigureCanvasTkAgg
    canvas = FigureCanvasTkAgg(fig, master=frame)
    canvas.draw()

    # Configura o fundo do widget canvas como transparente e sem borda
    canvas_widget = canvas.get_tk_widget()
    canvas_widget.config(bg="#004C70", highlightthickness=0)
    canvas_widget.pack(fill='both', expand=True)


def create_dashboard(parent_frame_a, parent_frame_b, parent_frame_c, total_occurrences,percentage_occurrences_with_error,taxon_count,percentage_taxons_with_error,top_taxon_errors):
    # Limpar o conteúdo anterior do frame
    for widget in parent_frame_a.winfo_children():
        widget.destroy()

    # Definir as cores para as labels
    text_color_white = 'white'
    dark_orange_color = '#FF8C00'  # Exemplo de cor laranja escuro

    # Criar label para "Total Occurrences"
    total_occurrences_label = customtkinter.CTkLabel(master=parent_frame_a,
                                                     text="Total Occurrences",
                                                     text_color=text_color_white,
                                                     font=(None,26),
                                                     corner_radius=8)
    total_occurrences_label.pack(pady=(10, 0))  # pack com um padding vertical apenas na parte superior

    # Criar label para exibir o valor de total_occurrences
    total_occurrences_value_label = customtkinter.CTkLabel(master=parent_frame_a,
                                                           text=str(total_occurrences),
                                                           text_color=dark_orange_color,
                                                           font=(None, 32),
                                                           corner_radius=8)
    total_occurrences_value_label.pack(pady=(5, 10))  # pack com padding vertical na parte superior e inferior
    add_pie_chart(parent_frame_a, percentage_occurrences_with_error)

    # Criar label para "Total taxon names"
    taxon_count_label = customtkinter.CTkLabel(master=parent_frame_b,
                                                     text="Total Taxon Names",
                                                     text_color=text_color_white,
                                                     font=(None, 26),
                                                     corner_radius=8)
    taxon_count_label.pack(pady=(10, 0))  # pack com um padding vertical apenas na parte superior

    # Criar label para exibir o valor de Total taxon names:
    taxon_count_value_label = customtkinter.CTkLabel(master=parent_frame_b,
                                                           text=str(taxon_count),
                                                           text_color=dark_orange_color,
                                                           font=(None, 32),
                                                           corner_radius=8)
    taxon_count_value_label.pack(pady=(5, 10))  # pack com padding vertical na parte superior e inferior
    add_pie_chart(parent_frame_b, percentage_taxons_with_error)

    # Criar label para "Total taxon names"
    top_taxon_errors_label = customtkinter.CTkLabel(master=parent_frame_c,
                                               text="Top 3 Taxonomic Ranks\nwith errors",
                                               text_color=text_color_white,
                                               font=(None, 26),
                                               corner_radius=8)
    top_taxon_errors_label.pack(pady=(10, 0))
    add_bar_graph(parent_frame_c, top_taxon_errors)


def create_layout():
    global tree, textbox, entry_input, entry_spreadsheet_name, corrected_spreadsheet_entry, entry_columns, frame2, frame_a, frame_b, frame_c

    log_file_path = "spreadsheet_log.txt"
    if os.path.exists(log_file_path):
        os.remove(log_file_path)

    #root = Tk()
    root = ThemedTk(theme="adapta")

    root.title("Pytaxon: a tool for detection and correction of taxonomic data error")
    root.geometry("1400x700")
    root.configure(bg='#002F3E')

    try:
        logo_image = Image.open('pytaxon/pytaxon_logo.png')
    except:
        logo_image = Image.open('_internal/pytaxon_logo.png')
    logo_photoimage = ImageTk.PhotoImage(logo_image.resize((315, 260), Image.Resampling.LANCZOS))
    logo_label = Label(master=root, image=logo_photoimage, bg='#002F3E')
    logo_label.image = logo_photoimage
    logo_label.place(relx=0.015, rely=0.05)

    frame_color = '#00546D'
    frame1 = ctk.CTkFrame(master=root, corner_radius=10, fg_color=frame_color)
    frame1.place(relx=0.015, rely=0.44, relwidth=0.23, relheight=0.51)

    label_input = ctk.CTkLabel(master=frame1, text="Input spreadsheet", fg_color=frame_color, text_color='white')
    label_input.place(relx=0.05, rely=0.1)
    entry_input = ctk.CTkEntry(master=frame1, placeholder_text="Search for your spreadsheet", fg_color="white")
    entry_input.place(relx=0.05, rely=0.2, relwidth=0.65)
    button_search = ctk.CTkButton(master=frame1, text="Search", fg_color="#004C70", hover_color="#0073A0",
                                  command=lambda: open_file(entry_input))
    button_search.place(relx=0.73, rely=0.2, relwidth=0.22)

    label_columns = ctk.CTkLabel(master=frame1, text="Column Names", fg_color=frame_color, text_color='white')
    label_columns.place(relx=0.05, rely=0.3)
    entry_columns = ctk.CTkEntry(master=frame1,
                                 placeholder_text="kingdom,phylum,class,order,family,genus,species,scientificName",
                                 fg_color="white")
    entry_columns.place(relx=0.05, rely=0.4, relwidth=0.9)

    label_source_id = ctk.CTkLabel(master=frame1, text="Source ID", fg_color=frame_color, text_color='white')
    label_source_id.place(relx=0.05, rely=0.5)
    option_menu_source_id = ctk.CTkOptionMenu(master=frame1, values=["1", "4", "11", "180"], fg_color="white")
    option_menu_source_id.place(relx=0.05, rely=0.6, relwidth=0.75)

    button_id_info = ctk.CTkButton(master=frame1, text="ID ?", fg_color="#004C70", hover_color="#0073A0",
                                   command=show_id_info)
    button_id_info.place(relx=0.84, rely=0.59, relwidth=0.12, relheight=0.09)

    label_check_spreadsheet = ctk.CTkLabel(master=frame1, text="Check Spreadsheet Name", fg_color=frame_color,
                                           text_color='white')
    label_check_spreadsheet.place(relx=0.05, rely=0.7)

    entry_spreadsheet_name = ctk.CTkEntry(master=frame1, placeholder_text="Spreadsheet name, no extension",
                                          fg_color="white")
    entry_spreadsheet_name.place(relx=0.05, rely=0.8, relwidth=0.65)
    button_run = ctk.CTkButton(master=frame1, text="Run", fg_color="#004C70", hover_color="#0073A0",
                               command=lambda: run_pytaxon(entry_input.get(), option_menu_source_id.get(),
                                                           entry_spreadsheet_name.get()))
    button_run.place(relx=0.73, rely=0.8, relwidth=0.22)

    frame2 = ctk.CTkFrame(master=root, corner_radius=10, fg_color=frame_color)
    frame2.place(relx=0.25, rely=0.05, relwidth=0.73, relheight=0.38)

    frame3 = ctk.CTkFrame(master=root, corner_radius=10, fg_color=frame_color)
    frame3.place(relx=0.25, rely=0.44, relwidth=0.36 * 0.65, relheight=0.51)

    # Configuração do Treeview dentro de frame3
    tree_frame3 = ttk.Frame(frame3)
    tree_frame3.place(relx=0.049, rely=0.1, relwidth=0.91, relheight=0.7)

    textbox = ctk.CTkTextbox(master=tree_frame3)
    textbox.pack(side='left', fill='both', expand=True)


    label_user_spreadsheet = Label(master=root, text="Running status", bg=frame_color, fg='white')
    label_user_spreadsheet.place(relx=0.26, rely=0.45)

    original_relwidth_frame4 = 0.36
    increase_factor = 0.35
    new_relwidth_frame4 = original_relwidth_frame4 * (1 + increase_factor)  # Aumenta a largura em 30%
    leftward_shift = original_relwidth_frame4 * increase_factor  # Calcula quanto o frame4 deve se mover para a esquerda

    # O novo relx é o antigo relx menos o deslocamento para a esquerda
    new_relx_frame4 = 0.62 - leftward_shift

    # Frame 4 - Adicionando barras de rolagem
    frame4 = ctk.CTkFrame(master=root, corner_radius=10, fg_color=frame_color)
    frame4.place(relx=new_relx_frame4, rely=0.44, relwidth=new_relwidth_frame4, relheight=0.51)

    label_check_spreadsheet = Label(master=root, text="Check Spreadsheet", bg=frame_color, fg='white')
    label_check_spreadsheet.place(relx=0.50, rely=0.45)

    # Definição da cor para os novos frames A, B, C
    new_frame_color = "#004C70"

    # Novas dimensões mantidas com um aumento de 15%
    frame_width = 0.28 * 1.15  # Aumento de 15% na largura
    frame_height = 0.8 * 1.15  # Aumento de 15% na altura

    # Ajustes para manter a proporção e o alinhamento após o aumento
    total_spacing = 1 - (3 * frame_width)  # Espaço total disponível após o redimensionamento
    spacing = total_spacing / 4  # Espaço entre os frames e nas laterais

    # Ajuste do parâmetro 'rely' para suspender ainda mais os frames
    rely_adjusted = 0.04  # Diminuindo ainda mais o 'rely'

    # Criar Frame A
    frame_a = ctk.CTkFrame(master=frame2, corner_radius=10, fg_color=new_frame_color)
    frame_a.place(relx=spacing, rely=rely_adjusted, relwidth=frame_width, relheight=frame_height)

    # Criar Frame B
    frame_b = ctk.CTkFrame(master=frame2, corner_radius=10, fg_color=new_frame_color)
    frame_b.place(relx=spacing * 2 + frame_width, rely=rely_adjusted, relwidth=frame_width, relheight=frame_height)

    # Criar Frame C
    frame_c = ctk.CTkFrame(master=frame2, corner_radius=10, fg_color=new_frame_color)
    frame_c.place(relx=spacing * 3 + 2 * frame_width, rely=rely_adjusted, relwidth=frame_width, relheight=frame_height)

    # Aumentar a largura do tree em 20%
    tree2_width_increase_factor = 1.09  # Aumentar a largura em 20%

    # Configuração do Treeview dentro de frame4
    tree_frame4 = ttk.Frame(frame4)
    tree_frame4.place(relx=0.015, rely=0.1, relwidth=0.89 * tree2_width_increase_factor, relheight=0.7)


    tree = ttk.Treeview(tree_frame4)
    tree.pack(side='left', fill='both', expand=True)

    scrollbar_vertical4 = ttk.Scrollbar(tree_frame4, orient='vertical', command=tree.yview)
    scrollbar_vertical4.pack(side='right', fill='y')

    scrollbar_horizontal4 = ttk.Scrollbar(tree_frame4, orient='horizontal', command=tree.xview)
    scrollbar_horizontal4.pack(side='bottom', fill='x')

    tree.configure(yscrollcommand=scrollbar_vertical4.set, xscrollcommand=scrollbar_horizontal4.set)
    tree.bind('<Double-1>', lambda event: on_double_click(event, tree, f"{entry_spreadsheet_name.get()}.xlsx"))


    corrected_spreadsheet_label = ctk.CTkLabel(master=frame4, text="Corrected spreadsheet name:", fg_color=frame_color,
                                               text_color='white')
    corrected_spreadsheet_label.place(relx=0.15, rely=0.82)

    corrected_spreadsheet_entry = ctk.CTkEntry(master=frame4, fg_color="white")
    corrected_spreadsheet_entry.place(relx=0.45, rely=0.82, relwidth=0.50)

    ignore_incertae_sedis_checkbox = ctk.CTkCheckBox(master=frame4, text="Ignore incertae sedis records",
                                                     fg_color=frame_color, text_color='white')
    ignore_incertae_sedis_checkbox.place(relx=0.11, rely=0.89)

    correct_button = ctk.CTkButton(master=frame4, text="Correct", fg_color="#004C70", hover_color="#0073A0",
                                   command=lambda: run_pytaxon_correct(entry_input.get(), corrected_spreadsheet_entry.get()))
    correct_button.place(relx=0.45, rely=0.90, relwidth=0.50)

    root.mainloop()


create_layout()
